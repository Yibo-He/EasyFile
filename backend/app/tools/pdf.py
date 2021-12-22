import os

'''
import fitz


def pdf2image(pdfPath,imgPath,zoom_x,zoom_y,rotation_angle):
    """
    # 将PDF转化为图片
    pdfPath pdf文件的路径
    imgPath 图像要保存的文件夹
    zoom_x x方向的缩放系数
    zoom_y y方向的缩放系数
    rotation_angle 旋转角度
    """
    # 打开PDF文件
    pdf = fitz.open(pdfPath)
    # 逐页读取PDF
    for pg in range(0, pdf.pageCount):
        page = pdf[pg]
        # 设置缩放和旋转系数
        trans = fitz.Matrix(zoom_x, zoom_y).prerotate(rotation_angle)
        pm = page.get_pixmap(matrix=trans, alpha=False)
        # 开始写图像
        pm.save(imgPath+str(pg)+".png")
    pdf.close()
'''

import camelot
import PyPDF2
import zipfile

def pdf2csv(pdfPath,x):
    tables = camelot.read_pdf(pdfPath,pages=x, line_scale=40)
    # for x in tables:
    #     camelot.plot(x, kind='text').show()

    #tables.export('test.csv', f='csv', compress=True) # json, excel, html, markdown, sqlite
    # print(len(tables))
    # tables[0].to_csv('test.csv',encoding='gbk') # to_json, to_excel, to_html, to_markdown, to_sqlite
    return tables


def split_pdf(document,lst):
    """
    切割指定的PDF文件
    """
    ret=[]
    for pages in lst:
        pdfw=PyPDF2.PdfFileWriter()
        for page in pages:
            pdfw.addPage(document.getPage(page))
        ret.append(pdfw)
    return ret


def pdf2chart(document, pages):
    '''
        extract the charts of the pages in the pdf document.
        Args:
            document: the return value of PyPDF2.PdfFileReader
            pages: list of number refering to page to process
        
        Return: a list of pd.DataFrame
    '''
    tables = []
    print([pages])
    pdf = split_pdf(document, [pages])
    print('line 65')
    tmp_path = document_path+'.tmp'
    with open(tmp_path,'wb') as f:
        pdf.write(f)
    tables += pdf2csv(tmp_path)
    os.remove(tmp_path)
    print(len(tables))
    return tables


from openpyxl import load_workbook
import openpyxl
import pandas as pd


def save_tables(tables, save_path):
    '''
        Save tables to a excel file.
        Args: 
            tables: 
            save_path: the path of the excel to save
    '''

    # workbook = openpyxl.Workbook()
    # workbook.save(filename=save_path)


    # writer = pd.ExcelWriter(save_path, engine='openpyxl')
    # book = load_workbook(save_path)
    # writer.book = book
    # for i, table in enumerate(tables):
    # table.to_excel(writer, sheet_name='chart_{}'.format(i), index=False)
    tables.export(save_path,f='excel')
    # writer.save()

def process(file_path,rules=['all'],format="csv",ext="csv",debug = False):
    tables = []
    filename="tmp.pdf"

    with open(filename, "wb") as f:
        f.write(open(file_path,'rb').read())
       # pdfs = split_pdf(document, rules)
    for i, x in enumerate(rules):
      #  with open('tmp' + str(i) + '.pdf', 'wb') as f:
        #    x.write(f)
        table = pdf2csv(filename,x)
        tables.append(table)
    os.remove(filename)
    zf = zipfile.ZipFile("output.zip","w")
    for i, table in enumerate(tables):
        table.export('part'+str(i)+"."+ext,f=format,compress=True)
        with zipfile.ZipFile('part'+str(i)+".zip","r") as zfin:
            zfin.extractall("./tmp")
        os.remove('part'+str(i)+'.'+'zip')
        # with open('output'+str(i)+'.'+'zip',"rb") as f:
        #     ret.append(f.read())
        for root, dirs, files in os.walk('./tmp'):
            for file in files:
                zf.write(os.path.join('./tmp',file),arcname='./part'+str(i)+'/'+file)
            for file in files:
                os.remove(os.path.join('./tmp', file))
        os.rmdir('./tmp')
        if debug:
            for j,x in enumerate(table):
                x.to_csv('test' + str(i)+','+str(j)+'.csv')
    zf.close()
    with open("output.zip","rb") as f:
        ret = f.read()
    os.remove("output.zip")
    return ret


if __name__ == "__main__":
    # pdf2image("test.pdf","test",10,10,0)
    filename = r"D:\Downloads\EasyFile\temp\None&0&test.pdf"
    x = open(filename,"rb")
    pdf = x.read()
    ret = process(pdf,rules=["2,3","1-2"],format='excel',ext="xlsx",debug=False)
    with open('out'+".zip","wb") as f:
        f.write(ret)